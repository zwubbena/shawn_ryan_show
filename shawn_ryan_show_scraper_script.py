# Scrape episode info from The Shawn Ryan Show.

# Pages: 1..55 of https://shawnryanshow.com/blogs/the-shawn-ryan-show
# Collects:
# - episode number
# - date (ISO format YYYY-MM-DD)
# - interviewee
# - roles / positions
# - full title
# - URL
# - source page

# Output:
# - shawn_ryan_episodes.xlsx, saved in this script's folder by default.
#
# How to use on macOS Terminal:
# 1. Optional but recommended:
#      cd insert-file-path-here/shawn_ryan_show
#      python3 -m venv .venv
#      source .venv/bin/activate
# 2. Run:
#      python3 shawn_ryan_show_scraper_script.py
#
# Optional flags:
#      python3 shawn_ryan_show_scraper_script.py --output episodes.xlsx
#      python3 shawn_ryan_show_scraper_script.py --first-page 1 --last-page 10
#      python3 shawn_ryan_show_scraper_script.py --extra-urls-file extra_urls.txt


# =========================
# Install / import packages
# =========================

from __future__ import annotations

import sys
import subprocess


def ensure_package(pkg_name: str, import_name: str | None = None):
    import_name = import_name or pkg_name
    try:
        return __import__(import_name)
    except ImportError:
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", pkg_name])
        except subprocess.CalledProcessError as exc:
            raise SystemExit(
                f"Could not install {pkg_name}. Run this first, then try again:\n"
                f"  {sys.executable} -m pip install requests beautifulsoup4 openpyxl"
            ) from exc
        return __import__(import_name)


def import_dependencies() -> None:
    global requests, BeautifulSoup, Workbook, Alignment, Font, PatternFill, get_column_letter

    requests = ensure_package("requests")
    bs4 = ensure_package("beautifulsoup4", "bs4")
    openpyxl = ensure_package("openpyxl")

    BeautifulSoup = bs4.BeautifulSoup
    Workbook = openpyxl.Workbook

    from openpyxl.styles import Alignment as OpenPyXLAlignment
    from openpyxl.styles import Font as OpenPyXLFont
    from openpyxl.styles import PatternFill as OpenPyXLPatternFill
    from openpyxl.utils import get_column_letter as openpyxl_get_column_letter

    Alignment = OpenPyXLAlignment
    Font = OpenPyXLFont
    PatternFill = OpenPyXLPatternFill
    get_column_letter = openpyxl_get_column_letter

import re
import time
import argparse
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Optional
from urllib.parse import urljoin, urlparse


# ==========
# Constants
# ==========

BASE_URL = "https://shawnryanshow.com"
BLOG_URL = f"{BASE_URL}/blogs/the-shawn-ryan-show"
FIRST_PAGE = 1
LAST_PAGE = 55
REQUEST_DELAY_SEC = 0.8
SCRIPT_DIR = Path(__file__).resolve().parent
OUTPUT_XLSX = SCRIPT_DIR / "shawn_ryan_episodes.xlsx"

DEFAULT_SEED_URLS = [
    "https://shawnryanshow.com/blogs/the-shawn-ryan-show/episode-001-mike-glover",
    "https://shawnryanshow.com/blogs/the-shawn-ryan-show/episode-002-travis-kennedy",
    "https://shawnryanshow.com/blogs/the-shawn-ryan-show/episode-003-don-bradley-a-k-a-headshot-don",
    "https://shawnryanshow.com/blogs/the-shawn-ryan-show/the-debrief-003-headshot-don",
    "https://shawnryanshow.com/blogs/the-shawn-ryan-show/episode-004-nick-kefalides",
    "https://shawnryanshow.com/blogs/the-shawn-ryan-show/the-debrief-004-nick-kefalides",
    "https://shawnryanshow.com/blogs/the-shawn-ryan-show/episode-005-mike-ritland-part-1",
    "https://shawnryanshow.com/blogs/the-shawn-ryan-show/episode-005-mike-ritland-part-2",
    "https://shawnryanshow.com/blogs/the-shawn-ryan-show/episode-006-travis-howze",
    "https://shawnryanshow.com/blogs/the-shawn-ryan-show/the-debrief-006-travis-howze",
    "https://shawnryanshow.com/blogs/the-shawn-ryan-show/shawn-ryan-show-007-drug-cartel-narcos-expert-ed-calderon",
    "https://shawnryanshow.com/blogs/the-shawn-ryan-show/the-debrief-007-ed-calderon",
    "https://shawnryanshow.com/blogs/the-shawn-ryan-show/shawn-ryan-show-008-retired-navy-seal-tried-for-murder-eddie-gallagher",
    "https://shawnryanshow.com/blogs/the-shawn-ryan-show/shawn-ryan-show-special-edition-wife-of-accused-navy-seal-war-criminal-andrea-gallagher",
    "https://shawnryanshow.com/blogs/the-shawn-ryan-show/us-mexico-border-crisis-2021-with-ed-calderon",
    "https://shawnryanshow.com/blogs/the-shawn-ryan-show/episode-005-tu-lam",
    "https://shawnryanshow.com/blogs/the-shawn-ryan-show/the-shawn-ryan-show-episode-010-marcus-capone",
    "https://shawnryanshow.com/blogs/the-shawn-ryan-show/shawn-ryan-show-011-the-blackwater-massacre",
    "https://shawnryanshow.com/blogs/the-shawn-ryan-show/shawn-ryan-show-012-luis-chaparro",
    "https://shawnryanshow.com/blogs/the-shawn-ryan-show/shawn-ryan-show-013-bernie-kerik",
    "https://shawnryanshow.com/blogs/the-shawn-ryan-show/shawn-ryan-show-014-alan-cooper",
    "https://shawnryanshow.com/blogs/the-shawn-ryan-show/shawn-ryan-show-015-seal-team-6-devgru-operator-dj-shipley",
    "https://shawnryanshow.com/blogs/the-shawn-ryan-show/the-shawn-ryan-show-016",
    "https://shawnryanshow.com/blogs/the-shawn-ryan-show/shawn-ryan-show-017-mikal-vega",
    "https://shawnryanshow.com/blogs/the-shawn-ryan-show/shawn-ryan-show-018-congressional-candidate-robby-starbuck",
    "https://shawnryanshow.com/blogs/the-shawn-ryan-show/shawn-ryan-show-the-china-influence-with-peter-schweizer",
    "https://shawnryanshow.com/blogs/the-shawn-ryan-show/shawn-ryan-show-020-navy-seal-justin-hughes",
    "https://shawnryanshow.com/blogs/the-shawn-ryan-show/shawn-ryan-show-inside-ukraine-with-mark-turner",
    "https://shawnryanshow.com/blogs/the-shawn-ryan-show/shawn-ryan-show-david-nino-rodriguez-heavyweight-boxing-champ",
    "https://shawnryanshow.com/blogs/the-shawn-ryan-show/shawn-ryan-show-inside-the-sinaloa-cartel-luis-chaparro",
    "https://shawnryanshow.com/blogs/the-shawn-ryan-show/shawn-ryans-psychedelic-experience",
    "https://shawnryanshow.com/blogs/the-shawn-ryan-show/eli-crane-navy-seal-sniper",
    "https://shawnryanshow.com/blogs/the-shawn-ryan-show/shawn-ryan-show-26-alex-epstein-and-the-energy-war",
    "https://shawnryanshow.com/blogs/the-shawn-ryan-show/shawn-ryan-show-26-rob-oneill-navy-seal",
    "https://shawnryanshow.com/blogs/the-shawn-ryan-show/shawn-ryan-show-28-nick-machine-lavery",
    "https://shawnryanshow.com/blogs/the-shawn-ryan-show/shawn-ryan-show-29-erik-prince-the-rise-and-fall-of-blackwater",
    "https://shawnryanshow.com/blogs/the-shawn-ryan-show/shawn-ryan-show-30-trevor-millar-psychedelic-provider",
    "https://shawnryanshow.com/blogs/the-shawn-ryan-show/shawn-ryan-show-031-marc-polymeropoulos-senior-cia-intelligence-officer",
    "https://shawnryanshow.com/blogs/the-shawn-ryan-show/shawn-ryan-show-32-jason-redman-navy-seal-talks-near-death-experience-and-seeking-redemption",
    "https://shawnryanshow.com/blogs/the-shawn-ryan-show/shawn-ryan-show-33-lily-tang-williams-i-fear-the-country-i-love-is-becoming-the-country-i-left",
    "https://shawnryanshow.com/blogs/the-shawn-ryan-show/shawn-ryan-show-34-cody-alford-marine-raider-marsoc-sniper-who-became-a-nomad",
    "https://shawnryanshow.com/blogs/the-shawn-ryan-show/dr-john-delony-the-mental-health-crisis",
    "https://shawnryanshow.com/blogs/the-shawn-ryan-show/lt-col-scott-mann-operation-pineapple-express",
    "https://shawnryanshow.com/blogs/the-shawn-ryan-show/037-machine-gun-preacher-sam-childers",
    "https://shawnryanshow.com/blogs/the-shawn-ryan-show/roger-reaves-pablo-escobar-and-the-medellin-cartels-1-drug-smuggler",
    "https://shawnryanshow.com/blogs/the-shawn-ryan-show/039-struggle-jennings",
    "https://shawnryanshow.com/blogs/the-shawn-ryan-show/40-mark-turner-ukraine-update",
    "https://shawnryanshow.com/blogs/the-shawn-ryan-show/41-kyle-morgan-delta-force-operator",
    "https://shawnryanshow.com/blogs/the-shawn-ryan-show/42-rob-luna-wealth-strategist-on-the-recession-and-inflation",
    "https://shawnryanshow.com/blogs/the-shawn-ryan-show/eddie-penney-seal-team-6-devgru-operator-043",
    "https://shawnryanshow.com/blogs/the-shawn-ryan-show/shawn-ryan-show-44-george-baker-agricultural-scientist",
]

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
}

MONTH_DATE_RE = re.compile(
    r"\b(?:January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{1,2},\s*\d{4}\b"
)

EP_NUM_PATTERNS = [
    re.compile(r"\bSRS\s*#?\s*(\d{1,4})\b", re.IGNORECASE),
    re.compile(r"\bepisode\s*[-#:]?\s*(\d{1,4})\b", re.IGNORECASE),
    re.compile(r"\bthe\s+debrief\s*[-#:]?\s*(\d{1,4})\b", re.IGNORECASE),
    re.compile(r"\bshawn\s+ryan\s+show\s*[-#:]?\s*(\d{1,4})\b", re.IGNORECASE),
    re.compile(r"(?:^|[\s/-])(\d{3})(?:[\s/-]|$)", re.IGNORECASE),
    re.compile(r"^\s*(\d{1,4})\b", re.IGNORECASE),
]
TITLE_GUEST_RE = re.compile(
    r"^\s*(?:SRS\s*#|episode|the\s+debrief|(?:the\s+)?shawn\s+ryan\s+show)\s*[-#:]?\s*\d+\s+(.*?)\s*-\s*(.*)$",
    re.IGNORECASE,
)
SPONSOR_STOP_RE = re.compile(r"shawn\s+ryan\s+show\s+sponsors", re.IGNORECASE)


# ==========
# Data model
# ==========

@dataclass
class Episode:
    source_page: int | str
    episode_number: Optional[str] = None
    date_iso: Optional[str] = None
    interviewee: Optional[str] = None
    roles: Optional[str] = None
    full_title: Optional[str] = None
    url: Optional[str] = None

    def as_row(self):
        episode_number = self.episode_number.zfill(3) if self.episode_number else ""
        return [
            episode_number,
            self.date_iso or "",
            self.interviewee or "",
            self.roles or "",
            self.full_title or "",
            self.url or "",
            self.source_page,
        ]


# ==================
# Utility functions
# ==================

def normalize_ws(text: str) -> str:
    return re.sub(r"\s+", " ", text or "").strip()


def title_from_url(url: str) -> str:
    slug = urlparse(url).path.rstrip("/").split("/")[-1]
    return normalize_ws(slug.replace("-", " ").replace("_", " ")).title()


def normalize_url(url: str) -> str:
    parsed = urlparse(urljoin(BASE_URL, url))
    return parsed._replace(query="", fragment="").geturl().rstrip("/")


def is_show_article_url(url: str) -> bool:
    parsed = urlparse(url)
    path = parsed.path.rstrip("/")
    prefix = "/blogs/the-shawn-ryan-show/"
    return path.startswith(prefix) and path != prefix.rstrip("/")


def to_iso_date(text: str) -> Optional[str]:
    text = normalize_ws(text)
    if not text:
        return None

    # Already ISO-like
    for fmt in ("%Y-%m-%d", "%Y-%m-%dT%H:%M:%S%z", "%Y-%m-%dT%H:%M:%S"):
        try:
            dt = datetime.strptime(text, fmt)
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            pass

    # Month day, year
    try:
        dt = datetime.strptime(text, "%B %d, %Y")
        return dt.strftime("%Y-%m-%d")
    except ValueError:
        return None

def extract_episode_number(text: str) -> Optional[str]:
    if not text:
        return None

    candidates = [
        normalize_ws(text),
        normalize_ws(text.replace("-", " ").replace("_", " ")),
    ]
    for candidate in candidates:
        for pattern in EP_NUM_PATTERNS:
            m = pattern.search(candidate)
            if m:
                return str(int(m.group(1)))
    return None


def extract_episode_number_from_url(url: str) -> Optional[str]:
    return extract_episode_number(title_from_url(url))

def parse_title_for_guest(title: str) -> tuple[Optional[str], Optional[str]]:
    """
    From:
      'SRS #296 Meg Appelgate - Why Parents Are Being Lied To About Teen “Treatment”'
    returns:
      ('296', 'Meg Appelgate')
    """
    title = normalize_ws(title)
    ep_num = extract_episode_number(title)
    m = TITLE_GUEST_RE.match(title)
    if m:
        guest = normalize_ws(m.group(1))
        return ep_num, guest
    return ep_num, None


# ===============
# HTTP / parsing
# ===============

def fetch(url: str, timeout: int = 30) -> str:
    last_error = None
    for attempt in range(3):
        try:
            resp = requests.get(url, headers=HEADERS, timeout=timeout)
            resp.raise_for_status()
            return resp.text
        except Exception as e:
            last_error = e
            print(f"  [warn] attempt {attempt+1} failed for {url}: {e}")
            time.sleep(2 * (attempt + 1))
    raise RuntimeError(f"Failed to fetch {url}: {last_error}")

def parse_listing_page(html: str, page_num: int) -> list[Episode]:
    """
    Collect episode links from a listing page.
    The Shawn Ryan site duplicates cards in parsed HTML, so we dedupe by URL.
    """
    soup = BeautifulSoup(html, "html.parser")
    episodes = []
    seen = set()

    anchors = soup.select('a[href*="/blogs/the-shawn-ryan-show/"]')
    for a in anchors:
        href = a.get("href", "").strip()
        if not href:
            continue

        full_url = normalize_url(href)
        if not is_show_article_url(full_url):
            continue
        if full_url in seen:
            continue
        seen.add(full_url)

        title_text = normalize_ws(a.get_text(" ", strip=True))
        fallback_title = title_from_url(full_url)
        ep_num, guest = parse_title_for_guest(title_text)
        ep_num = ep_num or extract_episode_number_from_url(full_url)

        if not ep_num and "shawn ryan show" not in title_text.lower():
            continue

        episodes.append(
            Episode(
                source_page=page_num,
                episode_number=ep_num,
                interviewee=guest,
                full_title=title_text or fallback_title,
                url=full_url,
            )
        )

    return episodes


def read_extra_urls(path: Path | None) -> list[str]:
    if not path:
        return []
    urls = []
    for line in path.expanduser().read_text().splitlines():
        line = line.strip()
        if line and not line.startswith("#"):
            urls.append(line)
    return urls


def seed_episodes(extra_urls: list[str], include_default_seeds: bool) -> list[Episode]:
    urls = []
    if include_default_seeds:
        urls.extend(DEFAULT_SEED_URLS)
    urls.extend(extra_urls)

    episodes = []
    seen = set()
    for url in urls:
        full_url = normalize_url(url)
        if full_url in seen or not is_show_article_url(full_url):
            continue
        seen.add(full_url)
        episodes.append(
            Episode(
                source_page="Seed",
                episode_number=extract_episode_number_from_url(full_url),
                full_title=title_from_url(full_url),
                url=full_url,
            )
        )
    return episodes

def parse_episode_page(html: str, fallback: Episode) -> Episode:
    soup = BeautifulSoup(html, "html.parser")

    # Title
    h1 = soup.find("h1")
    full_title = normalize_ws(h1.get_text(" ", strip=True)) if h1 else (fallback.full_title or "")

    ep_num, guest = parse_title_for_guest(full_title)
    ep_num = ep_num or extract_episode_number_from_url(fallback.url or "")
    if not guest:
        guest = fallback.interviewee

    # Date
    date_iso = None
    time_el = soup.find("time")
    if time_el:
        candidate = time_el.get("datetime") or time_el.get_text(" ", strip=True)
        date_iso = to_iso_date(candidate)

    if not date_iso:
        page_text = soup.get_text(" ", strip=True)
        m = MONTH_DATE_RE.search(page_text)
        if m:
            date_iso = to_iso_date(m.group(0))

    # Role summary / position
    paragraphs = []
    for p in soup.find_all("p"):
        txt = normalize_ws(p.get_text(" ", strip=True))
        if txt:
            paragraphs.append(txt)

    role_summary = None
    kept = []
    for p in paragraphs:
        if SPONSOR_STOP_RE.search(p):
            break
        kept.append(p)

    # Usually the first meaningful paragraph is the role/position summary.
    for p in kept:
        # skip trivial UI/share text
        low = p.lower()
        if low in {"share", "back", "close share copy link", "link"}:
            continue
        if len(p) < 40:
            continue
        role_summary = p
        break

    return Episode(
        source_page=fallback.source_page,
        episode_number=ep_num or fallback.episode_number,
        date_iso=date_iso,
        interviewee=guest,
        roles=role_summary,
        full_title=full_title or fallback.full_title,
        url=fallback.url,
    )


# ============
# Excel export
# ============

def write_xlsx(episodes: list[Episode], path: str | Path) -> None:
    path = Path(path).expanduser().resolve()
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Episodes"

    headers = [
        "Episode #",
        "Date (ISO)",
        "Interviewee",
        "Roles / Positions",
        "Full Title",
        "URL",
        "Source Page",
    ]
    ws.append(headers)

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="1F2937")
    header_align = Alignment(horizontal="center", vertical="center")

    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align

    body_font = Font(name="Arial")

    for ep in episodes:
        ws.append(ep.as_row())

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = [12, 14, 28, 60, 70, 85, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(path)


# =====
# Main
# =====

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Scrape Shawn Ryan Show episode info into an XLSX file."
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=OUTPUT_XLSX,
        help=f"Output XLSX path. Default: {OUTPUT_XLSX}",
    )
    parser.add_argument(
        "--first-page",
        type=int,
        default=FIRST_PAGE,
        help=f"First listing page to scrape. Default: {FIRST_PAGE}",
    )
    parser.add_argument(
        "--last-page",
        type=int,
        default=LAST_PAGE,
        help=f"Last listing page to scrape. Default: {LAST_PAGE}",
    )
    parser.add_argument(
        "--delay",
        type=float,
        default=REQUEST_DELAY_SEC,
        help=f"Seconds to wait between requests. Default: {REQUEST_DELAY_SEC}",
    )
    parser.add_argument(
        "--extra-urls-file",
        type=Path,
        default=None,
        help="Optional text file with one additional episode URL per line.",
    )
    parser.add_argument(
        "--no-default-seed-urls",
        action="store_true",
        help="Do not include the built-in early-episode seed URLs.",
    )
    return parser.parse_args()


def scrape_episodes(
    first_page: int,
    last_page: int,
    delay: float,
    extra_urls: list[str],
    include_default_seeds: bool,
) -> list[Episode]:
    if first_page < 1:
        raise ValueError("--first-page must be 1 or greater")
    if last_page < first_page:
        raise ValueError("--last-page must be greater than or equal to --first-page")
    if delay < 0:
        raise ValueError("--delay must be 0 or greater")

    all_listing_eps: list[Episode] = seed_episodes(
        extra_urls=extra_urls,
        include_default_seeds=include_default_seeds,
    )
    if all_listing_eps:
        print(f"Loaded {len(all_listing_eps)} seed episode URLs")

    print("Collecting episode links from listing pages...")
    for page in range(first_page, last_page + 1):
        url = BLOG_URL if page == 1 else f"{BLOG_URL}?page={page}"
        print(f"[listing {page:02d}/{last_page}] {url}")

        try:
            html = fetch(url)
            eps = parse_listing_page(html, page)
            print(f"  -> found {len(eps)} candidate episode links")
            all_listing_eps.extend(eps)
        except Exception as e:
            print(f"  [error] {e}")

        time.sleep(delay)

    # Dedupe by URL
    deduped_listing = []
    seen_urls = set()
    for ep in all_listing_eps:
        if not ep.url:
            continue
        if ep.url in seen_urls:
            continue
        seen_urls.add(ep.url)
        deduped_listing.append(ep)

    print(f"\nUnique episode links collected: {len(deduped_listing)}")

    if not deduped_listing:
        raise SystemExit(
            "No episodes were collected. The site structure may have changed."
        )

    # Visit each episode page
    final_episodes: list[Episode] = []
    print("\nScraping individual episode pages...")
    for idx, ep in enumerate(deduped_listing, start=1):
        print(f"[episode {idx:03d}/{len(deduped_listing)}] {ep.url}")
        try:
            html = fetch(ep.url)
            parsed = parse_episode_page(html, ep)
            final_episodes.append(parsed)
        except Exception as e:
            print(f"  [error] failed to parse {ep.url}: {e}")

        time.sleep(delay)

    # Final dedupe safeguard
    unique_final = []
    seen_keys = set()
    for ep in final_episodes:
        key = ep.url or (ep.full_title, ep.date_iso)
        if key in seen_keys:
            continue
        seen_keys.add(key)
        unique_final.append(ep)

    # Sort by episode number descending when possible
    def sort_key(ep: Episode):
        try:
            return int(ep.episode_number) if ep.episode_number else -1
        except ValueError:
            return -1

    unique_final.sort(key=sort_key, reverse=True)

    print(f"\nFinal unique episodes parsed: {len(unique_final)}")
    return unique_final


def main() -> None:
    args = parse_args()
    output_path = args.output.expanduser().resolve()

    import_dependencies()

    episodes = scrape_episodes(
        first_page=args.first_page,
        last_page=args.last_page,
        delay=args.delay,
        extra_urls=read_extra_urls(args.extra_urls_file),
        include_default_seeds=not args.no_default_seed_urls,
    )

    write_xlsx(episodes, output_path)
    print(f"Wrote {output_path}")


if __name__ == "__main__":
    main()
